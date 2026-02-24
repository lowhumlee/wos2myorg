"""
core.py — WoS MUV Affiliation Ingestion Tool
Core processing engine shared by CLI and Streamlit GUI.
Medical University of Varna · Research Information Systems
"""

from __future__ import annotations

import csv
import difflib
import io
import json
import logging
import os
import re
import sqlite3
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger("wos_muv.core")

# ─── Default Configuration ────────────────────────────────────────────────────

DEFAULT_CONFIG: dict = {
    "muv_affiliation_patterns": [
        "medical university varna",
        "med univ varna",
        "mu varna",
        "medical university of varna",
        "муварна",
        "медицинскиуниверситетварна",
    ],
    "fuzzy_threshold": 0.85,
    "interactive_mode": True,
    "allow_multi_org": True,
    "new_person_id_start": 9000,
    "output_dir": "output",
    "db_path": "staging.db",
}


def load_config(path: str = "config.json") -> dict:
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            cfg = json.load(f)
        return {**DEFAULT_CONFIG, **cfg}
    return DEFAULT_CONFIG.copy()


# ─── Name Normalization ───────────────────────────────────────────────────────

def strip_diacritics(text: str) -> str:
    """Remove diacritical marks, convert to ASCII."""
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_name(name: str) -> str:
    """
    Normalize author name for comparison.
    Output: lowercase ascii, punctuation stripped except comma, whitespace collapsed.
    """
    if not name:
        return ""
    name = strip_diacritics(name)
    name = name.lower()
    # Remove everything except letters, digits, commas, spaces
    name = re.sub(r"[^\w\s,]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def name_key(name: str) -> str:
    """Ultra-normalized key: letters only, lowercase ascii — for loose matching."""
    return re.sub(r"\s+", "", normalize_name(name).replace(",", ""))


def name_similarity(a: str, b: str) -> float:
    """Sequence-matcher similarity between two normalized names."""
    return difflib.SequenceMatcher(None, a, b).ratio()


# ─── C1 Field Parser ─────────────────────────────────────────────────────────

_C1_BLOCK = re.compile(r"\[([^\]]+)\]\s*([^[]+?)(?=\s*(?:\[|$))", re.DOTALL)


def parse_c1(c1: str) -> list[tuple[list[str], str]]:
    """
    Parse WoS C1 field into (author_list, affiliation_string) pairs.
    Handles both bracketed [A; B] Affil and bare affil (no brackets).
    """
    if not c1 or not c1.strip():
        return []
    results = []
    matches = _C1_BLOCK.findall(c1)
    if matches:
        for authors_raw, affil in matches:
            authors = [a.strip() for a in authors_raw.split(";") if a.strip()]
            affil = affil.strip().rstrip(";").strip()
            if affil:
                results.append((authors, affil))
    else:
        # No brackets — whole C1 is one affiliation for all authors
        results.append(([], c1.strip()))
    return results


# ─── MUV Affiliation Detection ───────────────────────────────────────────────

def is_muv_affiliation(affil: str, patterns: list[str]) -> bool:
    """
    Return True if affiliation string matches any MUV pattern.
    Strips diacritics, lowercases, and checks substring + fuzzy match.
    """
    if not affil:
        return False
    affil_norm = normalize_name(affil).replace(" ", "")
    for pat in patterns:
        pat_norm = normalize_name(pat).replace(" ", "")
        if pat_norm in affil_norm:
            return True
        # Short pattern: require ratio > 0.9
        if len(pat_norm) > 4 and name_similarity(pat_norm, affil_norm) > 0.90:
            return True
    return False


# ─── WoS Record Processing ───────────────────────────────────────────────────

def parse_wos_csv(content: str) -> list[dict]:
    """
    Parse WoS export CSV from string content.
    Strips whitespace from column names — WoS exports sometimes include
    leading/trailing spaces in headers (e.g. ' UT' instead of 'UT').
    """
    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for row in reader:
        # Skip None keys (caused by trailing delimiter in WoS exports)
        # Strip whitespace from all string keys and values
        clean = {}
        for k, v in row.items():
            if k is None:
                continue
            key = k.strip()
            if not key:
                continue
            val = v.strip() if isinstance(v, str) else (v or "")
            clean[key] = val
        if clean:
            rows.append(clean)
    return rows


def extract_muv_author_pairs(records: list[dict], cfg: dict) -> list[dict]:
    """
    For every WoS record, find authors with ≥1 MUV affiliation.
    Returns list of dicts:
      author_full, ut, muv_affils (list), all_affils (list)
    """
    patterns = cfg["muv_affiliation_patterns"]
    results = []

    for rec in records:
        ut = rec.get("UT", "").strip()
        af = rec.get("AF", "").strip()
        c1 = rec.get("C1", "").strip()
        if not ut or not af:
            continue

        af_authors = [a.strip() for a in af.split(";") if a.strip()]
        c1_blocks = parse_c1(c1)

        # author_name -> set of affiliations
        author_affils: dict[str, list[str]] = {a: [] for a in af_authors}

        for block_authors, affil in c1_blocks:
            if not block_authors:
                # Assign to all
                for a in af_authors:
                    author_affils[a].append(affil)
            else:
                # Match each C1 author to best AF author
                for c1_auth in block_authors:
                    c1_norm = normalize_name(c1_auth)
                    best, best_score = None, 0.0
                    for af_auth in af_authors:
                        s = name_similarity(c1_norm, normalize_name(af_auth))
                        if s > best_score:
                            best_score, best = s, af_auth
                    if best and best_score > 0.55:
                        if affil not in author_affils[best]:
                            author_affils[best].append(affil)

        for author, affils in author_affils.items():
            muv_affils = [a for a in affils if is_muv_affiliation(a, patterns)]
            if muv_affils:
                results.append({
                    "author_full": author,
                    "ut": ut,
                    "muv_affils": muv_affils,
                    "all_affils": affils,
                })

    logger.info("Found %d MUV author-document pairs.", len(results))
    return results


# ─── Existing Person Index ────────────────────────────────────────────────────

def build_person_index(content: str) -> tuple[dict[str, dict], int]:
    """
    Parse ResearcherAndDocument.csv content.
    Returns:
      index: normalized_name -> {PersonID, AuthorFullName, OrganizationID, ...}
      max_pid: highest numeric PersonID found
    """
    index: dict[str, dict] = {}
    max_pid = 0
    reader = csv.DictReader(io.StringIO(content))
    for row in reader:
        pid = row.get("PersonID", "").strip()
        first = row.get("FirstName", "").strip()
        last = row.get("LastName", "").strip()
        full = f"{last}, {first}" if last and first else (last or first)
        norm = normalize_name(full)
        org = row.get("OrganizationID", "").strip()
        if norm and pid:
            if norm not in index:
                index[norm] = {
                    "PersonID": pid,
                    "AuthorFullName": full,
                    "OrganizationID": org,
                }
        try:
            max_pid = max(max_pid, int(pid))
        except ValueError:
            pass
    logger.info("Built person index: %d unique persons (max PID=%d).", len(index), max_pid)
    return index, max_pid


def parse_org_hierarchy(content: str) -> list[dict]:
    """Parse OrganizationHierarchy.csv content."""
    reader = csv.DictReader(io.StringIO(content))
    orgs = []
    for row in reader:
        orgs.append({
            "OrganizationID": row.get("OrganizationID", "").strip(),
            "OrganizationName": row.get("OrganizationName", "").strip(),
            "ParentOrgaID": row.get("ParentOrgaID", "").strip(),
        })
    return orgs


# ─── Person Matching ──────────────────────────────────────────────────────────

def match_person(
    author_full: str,
    person_index: dict[str, dict],
    cfg: dict,
) -> tuple[str, list[tuple[str, dict, float]]]:
    """
    Match author name against existing person index.
    Returns (match_type, candidates):
      match_type: 'exact' | 'fuzzy' | 'new'
      candidates: sorted list of (norm_name, person_dict, score)
    """
    norm = normalize_name(author_full)
    threshold = float(cfg.get("fuzzy_threshold", 0.85))

    if norm in person_index:
        return "exact", [(norm, person_index[norm], 1.0)]

    candidates = []
    for existing_norm, person in person_index.items():
        score = name_similarity(norm, existing_norm)
        if score >= threshold:
            candidates.append((existing_norm, person, score))

    if candidates:
        candidates.sort(key=lambda x: -x[2])
        return "fuzzy", candidates

    return "new", []


# ─── Batch Processing (non-interactive) ──────────────────────────────────────

def batch_process(
    muv_pairs: list[dict],
    person_index: dict[str, dict],
    orgs: list[dict],
    cfg: dict,
    start_pid: int,
) -> dict:
    """
    Non-interactive processing. Auto-resolves exact matches.
    Returns result dict with:
      confirmed: list of output rows (exact matches)
      needs_review: list of candidates needing human decision
      new_persons: list of new person entries
    """
    confirmed: list[dict] = []
    needs_review: list[dict] = []
    seen_pairs: set[tuple] = set()
    new_persons: dict[str, dict] = {}  # norm -> person dict
    pid_counter = start_pid

    # Group by author to avoid re-processing same person
    author_groups: dict[str, list[dict]] = defaultdict(list)
    for pair in muv_pairs:
        norm = normalize_name(pair["author_full"])
        author_groups[norm].append(pair)

    for norm, pairs in author_groups.items():
        author_full = pairs[0]["author_full"]
        match_type, candidates = match_person(author_full, person_index, cfg)

        if match_type == "exact":
            person = candidates[0][1]
            pid = person["PersonID"]
            org_id = person.get("OrganizationID", "")
            for pair in pairs:
                key = (pid, pair["ut"], org_id)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)
                confirmed.append({
                    "PersonID": pid,
                    "AuthorFullName": person["AuthorFullName"],
                    "UT": pair["ut"],
                    "OrganizationID": org_id,
                    "muv_affils": pair["muv_affils"],
                    "match_type": "exact",
                    "norm": norm,
                })

        elif match_type == "fuzzy":
            top = candidates[0][1]
            for pair in pairs:
                needs_review.append({
                    "PersonID": "",
                    "AuthorFullName": author_full,
                    "UT": pair["ut"],
                    "muv_affils": pair["muv_affils"],
                    "match_type": "fuzzy",
                    "candidates": candidates,
                    "suggested_pid": top["PersonID"],
                    "suggested_name": top["AuthorFullName"],
                    "OrganizationID": top.get("OrganizationID", ""),
                    "norm": norm,
                })

        else:  # new
            if norm not in new_persons:
                new_persons[norm] = {
                    "PersonID": str(pid_counter),
                    "AuthorFullName": author_full,
                    "norm": norm,
                }
                pid_counter += 1
            person = new_persons[norm]
            for pair in pairs:
                needs_review.append({
                    "PersonID": person["PersonID"],
                    "AuthorFullName": author_full,
                    "UT": pair["ut"],
                    "muv_affils": pair["muv_affils"],
                    "match_type": "new",
                    "candidates": [],
                    "suggested_pid": person["PersonID"],
                    "suggested_name": author_full,
                    "OrganizationID": "",
                    "norm": norm,
                })

    return {
        "confirmed": confirmed,
        "needs_review": needs_review,
        "new_persons": new_persons,
        "next_pid": pid_counter,
    }


# ─── Output Generation ────────────────────────────────────────────────────────

def build_upload_csv(rows: list[dict], source_file: str) -> str:
    """Return upload-ready CSV as string."""
    out = io.StringIO()
    fields = ["PersonID", "AuthorFullName", "UT", "OrganizationID", "SourceFile", "Timestamp"]
    writer = csv.DictWriter(out, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    ts = datetime.now().isoformat(timespec="seconds")
    for row in rows:
        writer.writerow({
            "PersonID": row.get("PersonID", ""),
            "AuthorFullName": row.get("AuthorFullName", ""),
            "UT": row.get("UT", ""),
            "OrganizationID": row.get("OrganizationID", ""),
            "SourceFile": source_file,
            "Timestamp": ts,
        })
    return out.getvalue()


def build_audit_json(
    confirmed: list[dict],
    finalized: list[dict],
    rejected: list[dict],
    new_persons: list[dict],
) -> str:
    """Return audit log as JSON string."""
    audit = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "summary": {
            "exact_matches": len([r for r in confirmed if r.get("match_type") == "exact"]),
            "new_persons": len(new_persons),
            "finalized_records": len(finalized),
            "rejected_records": len(rejected),
        },
        "new_persons": [{"PersonID": p.get("PersonID"), "Name": p.get("AuthorFullName")} for p in new_persons],
        "finalized_records": finalized,
        "rejected_records": rejected,
    }
    return json.dumps(audit, indent=2, ensure_ascii=False)


def build_review_excel(
    needs_review: list[dict],
    orgs: list[dict],
) -> bytes:
    """Return Excel review workbook as bytes (requires openpyxl)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review Candidates"

    # ── Header styling
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
    GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
    ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")

    headers = [
        "PersonID", "AuthorFullName", "UT", "MUV Affiliations",
        "MatchType", "SuggestedPersonID", "SuggestedName",
        "OrganizationID",          # ← User fills this
        "APPROVED",                # YES / NO
        "NOTES",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    for row in needs_review:
        mt = row.get("match_type", "")
        fill = YELLOW_FILL if mt == "fuzzy" else (GREEN_FILL if mt == "new" else None)
        data = [
            row.get("PersonID", ""),
            row.get("AuthorFullName", ""),
            row.get("UT", ""),
            " | ".join(row.get("muv_affils", [])),
            mt,
            row.get("suggested_pid", ""),
            row.get("suggested_name", ""),
            row.get("OrganizationID", ""),  # blank for user
            "YES",
            "",
        ]
        ws.append(data)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    # Column widths
    widths = [12, 28, 22, 45, 10, 16, 28, 16, 10, 20]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Instructions sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        ["WoS MUV Affiliation Ingestion — Review Sheet"],
        [""],
        ["HOW TO USE THIS FILE:"],
        ["1. Review each row in the 'Review Candidates' sheet"],
        ["2. For FUZZY matches: verify that SuggestedPersonID is correct, or enter a different PersonID"],
        ["3. For NEW persons: confirm the PersonID or change it"],
        ["4. Fill in OrganizationID from the Organizations sheet"],
        ["5. Set APPROVED to YES to include, NO to exclude"],
        ["6. Save and re-import via the tool's batch import mode"],
        [""],
        ["MATCH TYPES:"],
        ["  new    = Author not found in existing dataset (yellow background)"],
        ["  fuzzy  = Possible match found above threshold (green background)"],
        ["  exact  = Already in dataset (handled automatically)"],
    ]
    for row in instructions:
        ws_info.append(row)
    ws_info["A1"].font = Font(bold=True, size=13, color="1F4E79")

    # ── Organizations reference sheet
    ws_org = wb.create_sheet("Organizations")
    ws_org.append(["OrganizationID", "OrganizationName", "ParentOrgaID"])
    for cell in ws_org[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    for org in orgs:
        indent = "    " if org["ParentOrgaID"] else ""
        ws_org.append([org["OrganizationID"], indent + org["OrganizationName"], org["ParentOrgaID"]])
    ws_org.column_dimensions["A"].width = 16
    ws_org.column_dimensions["B"].width = 55
    ws_org.column_dimensions["C"].width = 16
    ws_org.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Staging SQLite Database ──────────────────────────────────────────────────

class StagingDB:
    def __init__(self, path: str):
        self.path = path
        self.conn = sqlite3.connect(path, check_same_thread=False)
        self._init_schema()

    def _init_schema(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS persons (
                PersonID    TEXT PRIMARY KEY,
                FullName    TEXT,
                NormName    TEXT,
                IsNew       INTEGER DEFAULT 1,
                CreatedAt   TEXT
            );
            CREATE TABLE IF NOT EXISTS affiliations (
                ID          INTEGER PRIMARY KEY AUTOINCREMENT,
                PersonID    TEXT,
                UT          TEXT,
                OrgID       TEXT,
                RawAffil    TEXT,
                SourceFile  TEXT,
                Timestamp   TEXT,
                UNIQUE(PersonID, UT, OrgID)
            );
            CREATE TABLE IF NOT EXISTS decisions (
                ID          INTEGER PRIMARY KEY AUTOINCREMENT,
                PersonID    TEXT,
                DecisionType TEXT,
                Detail      TEXT,
                Timestamp   TEXT
            );
            CREATE TABLE IF NOT EXISTS rejected (
                ID          INTEGER PRIMARY KEY AUTOINCREMENT,
                AuthorName  TEXT,
                UT          TEXT,
                Reason      TEXT,
                Timestamp   TEXT
            );
        """)
        self.conn.commit()

    def upsert_person(self, pid: str, full_name: str, norm: str, is_new: bool = True):
        self.conn.execute(
            "INSERT OR IGNORE INTO persons VALUES (?,?,?,?,?)",
            (pid, full_name, norm, int(is_new), datetime.now().isoformat(timespec="seconds"))
        )
        self.conn.commit()

    def add_affiliation(self, pid: str, ut: str, org_id: str, raw_affil: str, source: str):
        try:
            self.conn.execute(
                "INSERT OR IGNORE INTO affiliations(PersonID,UT,OrgID,RawAffil,SourceFile,Timestamp) VALUES(?,?,?,?,?,?)",
                (pid, ut, org_id, raw_affil, source, datetime.now().isoformat(timespec="seconds"))
            )
            self.conn.commit()
        except Exception as e:
            logger.warning("DB affiliation insert error: %s", e)

    def log_decision(self, pid: str, dtype: str, detail: str):
        self.conn.execute(
            "INSERT INTO decisions(PersonID,DecisionType,Detail,Timestamp) VALUES(?,?,?,?)",
            (pid, dtype, detail, datetime.now().isoformat(timespec="seconds"))
        )
        self.conn.commit()

    def log_rejected(self, name: str, ut: str, reason: str):
        self.conn.execute(
            "INSERT INTO rejected(AuthorName,UT,Reason,Timestamp) VALUES(?,?,?,?)",
            (name, ut, reason, datetime.now().isoformat(timespec="seconds"))
        )
        self.conn.commit()

    def close(self):
        self.conn.close()
