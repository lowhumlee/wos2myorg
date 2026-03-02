"""
app.py — WoS MUV Affiliation Ingestion Tool · Streamlit GUI
Medical University of Varna · Research Information Systems

Run with:
  streamlit run app.py
"""

from __future__ import annotations

import csv
import io
import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import streamlit as st
import pandas as pd

# Make sure core.py is importable from same directory
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from core import (
    DEFAULT_CONFIG,
    normalize_name, name_similarity,
    build_person_index, parse_org_hierarchy, parse_wos_csv,
    extract_muv_author_pairs, match_person, batch_process,
    build_upload_csv, build_audit_json, build_review_excel,
    StagingDB,
)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── Page & Theme Setup ───────────────────────────────────────────────────────

st.set_page_config(
    page_title="WoS MUV Ingestion Tool",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.muv-header {
    background: linear-gradient(135deg, #0d2d4e 0%, #1a5276 60%, #2980b9 100%);
    border-radius: 14px;
    padding: 1.6rem 2.2rem 1.4rem;
    margin-bottom: 1.6rem;
    box-shadow: 0 4px 20px rgba(0,0,0,0.18);
}
.muv-header h1 { color: #ffffff; margin: 0; font-size: 1.75rem; letter-spacing: -0.02em; }
.muv-header .sub { color: #a8d4f5; margin: 0.35rem 0 0; font-size: 0.92rem; }

.metric-grid { display: flex; gap: 1rem; margin: 1rem 0; flex-wrap: wrap; }
.metric-card {
    flex: 1 1 140px;
    background: #fff;
    border: 1px solid #d0dde8;
    border-radius: 12px;
    padding: 1.1rem 1.2rem;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}
.metric-card .num { font-size: 2.4rem; font-weight: 800; line-height: 1; }
.metric-card .lbl { font-size: 0.78rem; color: #6b7a8d; margin-top: 0.3rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.04em; }
.num-blue { color: #1a5276; }
.num-green { color: #1e8449; }
.num-orange { color: #d35400; }
.num-yellow { color: #9a7d0a; }

.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.78rem;
    font-weight: 700;
    letter-spacing: 0.03em;
}
.badge-new      { background: #d5f5e3; color: #1e8449; }
.badge-exact    { background: #d6eaf8; color: #1a5276; }
.badge-fuzzy    { background: #fef9e7; color: #9a7d0a; border: 1px solid #f9e79f; }
.badge-initial  { background: #fff3cd; color: #856404; border: 1px solid #ffc107; }

.sec-head {
    background: #eaf2fb;
    border-left: 5px solid #1a5276;
    padding: 0.55rem 1rem;
    border-radius: 0 8px 8px 0;
    margin: 1.2rem 0 0.6rem;
    font-weight: 700;
    color: #1a3a5c;
    font-size: 1rem;
}

.chip {
    display: inline-block;
    background: #eaf2fb;
    border: 1px solid #aed6f1;
    color: #1a5276;
    padding: 3px 10px;
    border-radius: 16px;
    font-size: 0.8rem;
    margin: 2px;
}

div[data-testid="stFileUploader"] {
    border: 2px dashed #aed6f1;
    border-radius: 10px;
    padding: 0.4rem;
}

div.stButton > button { border-radius: 8px; font-weight: 600; }
div.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #1a5276, #2980b9);
    border: none;
    color: white;
}

div[data-testid="stExpander"] {
    border: 1px solid #d0dde8 !important;
    border-radius: 10px !important;
    margin-bottom: 0.5rem;
}

div.stDownloadButton > button {
    border-radius: 8px;
    font-weight: 600;
    background: #1e8449;
    color: white;
    border: none;
}

.sibling-note {
    background: #f0f4ff;
    border: 1px solid #c5d0f0;
    border-radius: 6px;
    padding: 4px 10px;
    font-size: 0.8rem;
    color: #3a4a8a;
    margin-bottom: 6px;
    display: inline-block;
}

/* ── Decision status pills (shown in minimized expander label) ── */
.status-done {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 12px;
    font-size: 0.75rem;
    font-weight: 700;
    background: #d4edda;
    color: #155724;
    margin-left: 6px;
}
.status-rejected {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 12px;
    font-size: 0.75rem;
    font-weight: 700;
    background: #f8d7da;
    color: #721c24;
    margin-left: 6px;
}
</style>
""", unsafe_allow_html=True)

# ─── Session State Init ───────────────────────────────────────────────────────

_DEFAULTS = {
    "cfg": DEFAULT_CONFIG.copy(),
    "person_index": {},
    "max_pid": 0,
    "orgs": [],
    "wos_records": [],
    "muv_pairs": [],
    "batch_result": None,
    "decisions": {},
    "output_rows": [],
    "rejected_rows": [],
    "processed": False,
    "finalized": False,
    "source_file": "unknown.csv",
}

for k, v in _DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


def reset_state():
    for k, v in _DEFAULTS.items():
        st.session_state[k] = v if not callable(v) else v()


# ─── Sidebar ─────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("### ⚙️ Configuration")

    cfg = st.session_state.cfg
    cfg["fuzzy_threshold"] = st.slider(
        "Fuzzy match threshold", 0.5, 1.0,
        float(cfg.get("fuzzy_threshold", 0.85)), 0.01,
        help="Minimum name similarity to flag as possible match (0 = everything, 1 = exact only)"
    )
    cfg["allow_multi_org"] = st.checkbox(
        "Allow multiple organizations per author",
        value=bool(cfg.get("allow_multi_org", True))
    )
    st.caption(
        "New PersonIDs are assigned automatically as max(existing) + 1"
    )

    st.markdown("---")
    st.markdown("### 🔍 MUV Affiliation Patterns")
    pat_text = st.text_area(
        "Patterns (one per line, case-insensitive)",
        value="\n".join(cfg.get("muv_affiliation_patterns", [])),
        height=160,
        help="Substrings that identify a MUV affiliation in WoS C1 field"
    )
    cfg["muv_affiliation_patterns"] = [p.strip() for p in pat_text.splitlines() if p.strip()]

    st.markdown("---")
    if st.button("🔄 Reset All", use_container_width=True):
        reset_state()
        st.rerun()

    st.markdown("---")
    st.caption("WoS MUV Ingestion Tool · v2.1")
    st.caption("Medical University of Varna")

# ─── Header ──────────────────────────────────────────────────────────────────

st.markdown("""
<div class="muv-header">
  <h1>🔬 WoS My Organization — Affiliation Ingestion Tool</h1>
  <div class="sub">Medical University of Varna · Bibliometric Data Curation Workflow</div>
</div>
""", unsafe_allow_html=True)

# ─── Live person search helper ───────────────────────────────────────────────

def search_person_index(query: str, person_index: list, max_results: int = 6) -> list:
    """
    Search the person index for names matching `query`.
    Returns a ranked list of (score, person_dict) tuples.
    Handles both surname-first ("Chaushev, Borislav") and
    given-first ("Borislav Chaushev") queries.
    """
    from core import normalize_name, name_similarity
    if not query or len(query) < 2:
        return []

    q = normalize_name(query)
    # Also try reversing "Firstname Lastname" → "Lastname, Firstname"
    q_parts = q.split()
    q_reversed = f"{q_parts[-1]} {' '.join(q_parts[:-1])}" if len(q_parts) > 1 else q

    results = []
    for p in person_index:
        name = p["NormName"]
        # Score against both orderings
        score = max(
            name_similarity(q, name),
            name_similarity(q_reversed, name),
        )
        # Boost if the query is a substring of the name or vice versa
        if q in name or any(part in name for part in q.split() if len(part) > 2):
            score = max(score, 0.5)
        if score >= 0.3:
            results.append((score, p))

    results.sort(key=lambda x: -x[0])
    return results[:max_results]


# ─── Tabs ─────────────────────────────────────────────────────────────────────

tab_load, tab_review, tab_output, tab_stats, tab_help = st.tabs([
    "📂 1 · Load Files",
    "🔎 2 · Review & Resolve",
    "📤 3 · Export Output",
    "📊 4 · Statistics",
    "❓ Help",
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — LOAD FILES
# ══════════════════════════════════════════════════════════════════════════════

with tab_load:
    st.markdown('<div class="sec-head">Upload Input Files</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("**📋 WoS Export (new records)**")
        wos_file = st.file_uploader("WoS CSV (input*.csv)", type=["csv", "txt"], key="wos_up")
        if wos_file:
            st.success(f"✓ {wos_file.name}")
            st.session_state.source_file = wos_file.name

    with col2:
        st.markdown("**👥 Existing Researchers**")
        res_file = st.file_uploader("ResearcherAndDocument.csv", type=["csv"], key="res_up")
        if res_file:
            st.success(f"✓ {res_file.name}")

    with col3:
        st.markdown("**🏛️ Organization Hierarchy**")
        org_file = st.file_uploader("OrganizationHierarchy.csv", type=["csv"], key="org_up")
        if org_file:
            st.success(f"✓ {org_file.name}")

    st.markdown("---")

    if res_file:
        res_file.seek(0)
        df_preview = list(csv.DictReader(io.StringIO(res_file.read().decode("utf-8-sig"))))
        res_file.seek(0)
        with st.expander(f"Preview: Researchers ({len(df_preview)} rows)", expanded=False):
            st.dataframe(pd.DataFrame(df_preview).head(15), use_container_width=True)

    if org_file:
        org_file.seek(0)
        df_org_preview = list(csv.DictReader(io.StringIO(org_file.read().decode("utf-8-sig"))))
        org_file.seek(0)
        with st.expander(f"Preview: Organizations ({len(df_org_preview)} orgs)", expanded=False):
            st.dataframe(pd.DataFrame(df_org_preview), use_container_width=True)

    if wos_file:
        wos_file.seek(0)
        wos_preview_raw = wos_file.read().decode("utf-8-sig")
        wos_file.seek(0)
        wos_preview = list(csv.DictReader(io.StringIO(wos_preview_raw)))
        with st.expander(f"Preview: WoS Records ({len(wos_preview)} records)", expanded=False):
            if wos_preview:
                df_wos_prev = pd.DataFrame(wos_preview)
                df_wos_prev.columns = [
                    c.strip() if c is not None else "__EMPTY__"
                    for c in df_wos_prev.columns
                ]
                df_wos_prev = df_wos_prev[[c for c in df_wos_prev.columns if c != "__EMPTY__"]]
                preferred = [c for c in ["UT", "AF"] if c in df_wos_prev.columns]
                st.caption(f"Detected columns: {list(df_wos_prev.columns)}")
                if preferred:
                    st.dataframe(df_wos_prev[preferred].head(10), use_container_width=True)
                else:
                    st.dataframe(df_wos_prev.head(10), use_container_width=True)
            else:
                st.warning("No rows found in WoS file.")

    st.markdown("---")

    proc_btn = st.button(
        "🚀  Detect MUV Authors",
        type="primary",
        disabled=not (wos_file and res_file and org_file),
        use_container_width=True,
    )

    if proc_btn:
        cfg = st.session_state.cfg

        with st.spinner("Parsing files and detecting MUV-affiliated authors…"):
            # Read all file contents once
            res_file.seek(0)
            res_content = res_file.read().decode("utf-8-sig")
            org_file.seek(0)
            org_content = org_file.read().decode("utf-8-sig")
            wos_file.seek(0)
            wos_content = wos_file.read().decode("utf-8-sig")

            person_index, max_pid, existing_pairs = build_person_index(res_content)
            orgs                  = parse_org_hierarchy(org_content)
            records               = parse_wos_csv(wos_content)
            muv_pairs             = extract_muv_author_pairs(records, cfg)

            # Always start above the highest existing PersonID in the master file
            # so new IDs never clash with previously uploaded records
            start_pid = max_pid + 1

            batch_result = batch_process(
                muv_pairs, person_index, orgs, cfg, start_pid,
                researcher_csv_content=res_content,
                existing_pairs=existing_pairs,
            )

            # Build decisions dict for interactive review
            decisions_by_norm: dict[str, dict] = {}
            for item in batch_result["needs_review"]:
                norm = item["norm"]
                if norm not in decisions_by_norm:
                    # Use org IDs from the matched master record as the pre-filled default
                    default_org_ids = item.get("suggested_org_ids") or                                        ([item["OrganizationID"]] if item.get("OrganizationID") else [""])
                    # New persons start un-approved (require explicit sign-off).
                    # Fuzzy / initial_expansion matches start pre-approved
                    # (user just needs to confirm or reject the suggested identity).
                    mt_item = item.get("match_type", "new")
                    auto_approve = mt_item in ("initial_expansion", "fuzzy", "resolved")
                    decisions_by_norm[norm] = {
                        **item,
                        "org_ids":       default_org_ids,
                        "resolved_pid":  item.get("suggested_pid", ""),
                        "resolved_name": item.get("suggested_name",
                                                   item.get("AuthorFullName", "")),
                        "approved": auto_approve,
                    }

            st.session_state.person_index  = person_index
            st.session_state.max_pid       = max_pid
            st.session_state.orgs          = orgs
            st.session_state.wos_records   = records
            st.session_state.muv_pairs     = muv_pairs
            st.session_state.batch_result  = batch_result
            st.session_state.decisions     = decisions_by_norm
            st.session_state.existing_pairs = existing_pairs
            st.session_state.processed     = True
            st.session_state.finalized     = False
            st.session_state.output_rows   = []

        confirmed = batch_result["confirmed"]
        review    = batch_result["needs_review"]
        new_p     = batch_result["new_persons"]

        already_up_all    = batch_result.get("already_uploaded", [])
        n_exact           = len([r for r in confirmed   if r.get("match_type") == "exact"])
        n_new             = len([r for r in review      if r.get("match_type") == "new"])
        n_fuzzy           = len([r for r in review      if r.get("match_type") == "fuzzy"])
        n_initial         = len([r for r in review      if r.get("match_type") == "initial_expansion"])
        n_uploaded_exact  = len([r for r in already_up_all if r.get("match_type") != "probable_duplicate"])
        n_prob_dup        = len([r for r in already_up_all if r.get("match_type") == "probable_duplicate"])
        n_uploaded        = len(already_up_all)

        st.markdown(f"""
<div class="metric-grid">
  <div class="metric-card"><div class="num num-blue">{len(muv_pairs)}</div><div class="lbl">MUV Pairs Found</div></div>
  <div class="metric-card"><div class="num num-blue">{n_exact}</div><div class="lbl">Auto-Confirmed (exact)</div></div>
  <div class="metric-card"><div class="num num-green">{len(new_p)}</div><div class="lbl">New Persons Staged</div></div>
  <div class="metric-card"><div class="num num-yellow">{n_initial}</div><div class="lbl">Initial-Expansion Matches</div></div>
  <div class="metric-card"><div class="num num-yellow">{n_fuzzy}</div><div class="lbl">Fuzzy / Ambiguous</div></div>
  <div class="metric-card"><div class="num num-orange">{len(review)}</div><div class="lbl">Needs Review</div></div>
  <div class="metric-card"><div class="num" style="color:#888">{n_uploaded_exact}</div><div class="lbl">Already in MyOrg</div></div>
  <div class="metric-card"><div class="num" style="color:#e67e22">{n_prob_dup}</div><div class="lbl">Probable Duplicates</div></div>
</div>
""", unsafe_allow_html=True)
        if n_uploaded_exact:
            st.info(f"ℹ️ **{n_uploaded_exact} pair(s) skipped** — already present in ResearcherAndDocument.csv "
                    f"and do not need re-uploading.")
        if n_prob_dup:
            st.warning(
                f"⚠️ **{n_prob_dup} probable duplicate(s)** — high-confidence name matches whose "
                f"(PersonID, UT) combination already exists in MyOrg. "
                f"These have been skipped automatically. Review them in the **Statistics tab**."
            )

        if len(review) > 0:
            st.info(f"➡️ **{len(review)} entries need your decision.** Go to Tab 2 to review.")
        else:
            st.success("✅ All authors matched automatically. Go to Tab 3 to export.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — REVIEW & RESOLVE
# ══════════════════════════════════════════════════════════════════════════════

with tab_review:
    if not st.session_state.processed:
        st.info("⬅️ Please load and process data in **Tab 1** first.")
    else:
        batch_result = st.session_state.batch_result
        decisions    = st.session_state.decisions
        orgs         = st.session_state.orgs
        cfg          = st.session_state.cfg

        confirmed_auto = batch_result["confirmed"]
        needs_review   = batch_result["needs_review"]

        # ── Filters ──────────────────────────────────────────────────────────
        fcol1, fcol2, fcol3 = st.columns([2, 2, 1])
        with fcol1:
            ftype = st.selectbox("Filter", [
                "Pending decisions only",
                "All needing review",
                "New persons only",
                "Fuzzy matches only",
                "Initial-expansion matches only",
                "Approved",
                "Rejected",
            ])
        with fcol2:
            fsearch = st.text_input("Search author name", "")
        with fcol3:
            st.markdown("<br>", unsafe_allow_html=True)
            show_exact = st.checkbox("Show auto-confirmed", value=False)

        # ── Org dropdown helpers ──────────────────────────────────────────────
        org_map    = {f"[{o['OrganizationID']}] {o['OrganizationName']}": o["OrganizationID"]
                      for o in orgs}
        org_labels = ["— none / skip —"] + list(org_map.keys())

        def label_for_org(oid: str) -> str:
            for lbl, v in org_map.items():
                if v == oid:
                    return lbl
            return org_labels[0]

        # ── Auto-confirmed section ────────────────────────────────────────────
        if show_exact and confirmed_auto:
            st.markdown('<div class="sec-head">✅ Auto-Confirmed (Exact Matches)</div>',
                        unsafe_allow_html=True)
            df_conf = pd.DataFrame([{
                "PersonID": r["PersonID"],
                "Name": r["AuthorFullName"],
                "UT": r["UT"],
                "OrgID": r["OrganizationID"],
            } for r in confirmed_auto])
            st.dataframe(df_conf, use_container_width=True, height=200)

        # ── Needs-review section ──────────────────────────────────────────────
        st.markdown(
            f'<div class="sec-head">🔍 Needs Human Decision ({len(needs_review)} entries)</div>',
            unsafe_allow_html=True,
        )

        # Group by SiblingGroup first, then by norm within each group
        # so variants of the same author appear adjacent in the UI
        by_norm: dict[str, list] = defaultdict(list)
        for item in needs_review:
            norm = item["norm"]
            mt   = item["match_type"]
            if ftype == "New persons only"                 and mt != "new":               continue
            if ftype == "Fuzzy matches only"               and mt != "fuzzy":             continue
            if ftype == "Initial-expansion matches only"   and mt != "initial_expansion": continue
            if fsearch and fsearch.lower() not in item["AuthorFullName"].lower():         continue
            by_norm[norm].append(item)

        if not by_norm:
            st.success("✅ No entries match the current filter.")
        else:
            # Sort norms by SiblingGroup so siblings appear together
            def sort_key(norm_items):
                return norm_items[1][0].get("SiblingGroup", norm_items[1][0]["AuthorFullName"])

            sorted_items = sorted(by_norm.items(), key=sort_key)

            # Progress summary — only count items where user has explicitly decided
            total_items   = len(sorted_items)
            decided_items = sum(
                1 for norm, _ in sorted_items
                if decisions.get(norm, {}).get("user_decided", False)
            )
            approved_items  = sum(
                1 for norm, _ in sorted_items
                if decisions.get(norm, {}).get("user_decided", False)
                and decisions.get(norm, {}).get("approved", True)
            )
            rejected_items  = decided_items - approved_items
            pending_items   = total_items - decided_items

            st.markdown(
                f'<div style="margin:0.5rem 0 1rem;font-size:0.88rem;color:#555;">'
                f'<b>{total_items}</b> entries &nbsp;·&nbsp; '
                f'<span style="color:#1e8449"><b>{approved_items}</b> approved</span> &nbsp;·&nbsp; '
                f'<span style="color:#c0392b"><b>{rejected_items}</b> rejected</span> &nbsp;·&nbsp; '
                f'<span style="color:#d35400"><b>{pending_items}</b> pending</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

            prev_sibling_group = None

            for norm, items in sorted_items:
                first  = items[0]
                mt     = first["match_type"]
                author = first["AuthorFullName"]
                sibling_group = first.get("SiblingGroup", author)

                # ── Determine current decision status ─────────────────────────
                dec_current  = decisions.get(norm)
                # user_decided = True only after the user has explicitly ticked
                # the approve/reject checkbox (not just from pre-population at load time)
                is_decided   = dec_current is not None and dec_current.get("user_decided", False)
                is_approved  = is_decided and dec_current.get("approved", True)
                is_rejected  = is_decided and not dec_current.get("approved", True)

                # Apply filter
                if ftype == "Pending decisions only"  and is_decided:   continue
                if ftype == "Approved"                and not is_approved: continue
                if ftype == "Rejected"                and not is_rejected: continue

                # ── Sibling group divider ─────────────────────────────────────
                if sibling_group != prev_sibling_group:
                    prev_sibling_group = sibling_group
                    sibling_norms = [
                        n for n, its in sorted_items
                        if its[0].get("SiblingGroup", its[0]["AuthorFullName"]) == sibling_group
                    ]
                    if len(sibling_norms) > 1:
                        sibling_names = [by_norm[n][0]["AuthorFullName"]
                                         for n in sibling_norms if n in by_norm]
                        st.markdown(
                            f'<div class="sibling-note">🔗 Sibling group: '
                            f'{" · ".join(sibling_names)}</div>',
                            unsafe_allow_html=True,
                        )

                # ── Match type badge ──────────────────────────────────────────
                badge_html = {
                    "new":                '<span class="badge badge-new">🆕 NEW PERSON</span>',
                    "fuzzy":              '<span class="badge badge-fuzzy">⚠ AMBIGUOUS MATCH</span>',
                    "initial_expansion":  '<span class="badge badge-initial">🔤 INITIAL MATCH — please confirm</span>',
                }.get(mt, "")

                uts_str = ", ".join(i["UT"] for i in items)

                # ── Build expander label with status suffix ───────────────────
                if is_rejected:
                    label = f"❌  {author}  —  {len(items)} document(s)  · REJECTED"
                elif is_approved and is_decided:
                    resolved = dec_current.get("resolved_name", author)
                    org_ids  = [o for o in dec_current.get("org_ids", []) if o]
                    org_str  = ", ".join(org_ids) if org_ids else "no org"
                    label    = f"✅  {author}  →  {resolved}  [{org_str}]"
                elif not is_decided and mt == "initial_expansion":
                    label    = f"🔤  {author}  —  confirm match  ({len(items)} doc)"
                elif not is_decided and mt in ("fuzzy",):
                    label    = f"⚠  {author}  —  review needed  ({len(items)} doc)"
                elif not is_decided and mt == "new":
                    label    = f"🆕  {author}  —  new person, needs org  ({len(items)} doc)"
                else:
                    label    = f"⏳  {author}  —  {len(items)} document(s)"

                # Expand when not yet decided by user
                should_expand = not is_decided

                with st.expander(label, expanded=should_expand):
                    st.markdown(badge_html, unsafe_allow_html=True)

                    # MUV affiliation chips
                    all_muv   = []
                    for it in items:
                        all_muv.extend(it["muv_affils"])
                    unique_muv = list(dict.fromkeys(all_muv))
                    chips_html = " ".join(f'<span class="chip">{a}</span>' for a in unique_muv[:4])
                    st.markdown(f"<small><b>MUV affiliations:</b> {chips_html}</small>",
                                unsafe_allow_html=True)
                    st.markdown(f"<small><b>Documents:</b> {uts_str}</small>",
                                unsafe_allow_html=True)

                    # ── Identity decision ─────────────────────────────────────
                    dec = decisions.get(norm, {
                        "resolved_pid":  first.get("suggested_pid", ""),
                        "resolved_name": first.get("AuthorFullName", ""),
                        "org_ids":       [""],
                        "approved":      mt != "new",
                        "user_decided":  False,
                    })

                    id_col1, id_col2 = st.columns(2)

                    with id_col1:
                        # Both fuzzy AND initial_expansion show a candidate picker
                        if mt in ("fuzzy", "initial_expansion") and first.get("candidates"):
                            cands = first["candidates"]
                            cand_labels = [
                                f"[{p['PersonID']}] {p['AuthorFullName']} ({s:.2f})"
                                for _, p, s in cands
                            ]
                            cand_labels.append("➕ Create as NEW PERSON")

                            # Pre-select the top candidate
                            default_idx = 0

                            choice = st.selectbox(
                                f"Identity for {author}",
                                cand_labels,
                                index=default_idx,
                                key=f"identity_{norm}",
                            )
                            if "NEW PERSON" in choice:
                                dec["resolved_pid"]  = first.get("suggested_pid", "")
                                dec["resolved_name"] = author
                                dec["match_type"]    = "new"
                                # Clear pre-filled orgs so user assigns manually
                                dec["org_ids"] = [""]
                            else:
                                idx = cand_labels.index(choice)
                                _, chosen_person, _ = cands[idx]
                                dec["resolved_pid"]  = chosen_person["PersonID"]
                                dec["resolved_name"] = chosen_person["AuthorFullName"]
                                dec["match_type"]    = "resolved"
                                # Pre-fill org IDs from the chosen master record
                                dec["org_ids"] = chosen_person.get("OrganizationIDs") or                                     ([chosen_person["OrganizationID"]]
                                     if chosen_person.get("OrganizationID") else [""])
                        else:
                            # ── Identity picker for new persons ──────────────
                            # Build candidate list: live search + "create new" option
                            search_query = st.text_input(
                                "Search existing researchers",
                                value=dec.get("search_query", author),
                                key=f"search_{norm}",
                                help="Type a name to find existing persons — then select below",
                            )
                            dec["search_query"] = search_query

                            live_hits = search_person_index(
                                search_query, list(st.session_state.person_index)
                            )

                            # Build selectbox options:
                            # option 0 = "➕ Create as NEW PERSON"
                            # options 1..N = matched existing persons
                            NEW_PERSON_LABEL = f"➕ Create as NEW PERSON  (ID will be {first.get('suggested_pid', 'auto')})"
                            hit_options = [NEW_PERSON_LABEL]
                            hit_map: dict[str, dict] = {}  # label → person dict
                            for hit_score, hit_person in live_hits:
                                lbl = (
                                    f"[{hit_person['PersonID']}] {hit_person['AuthorFullName']}"
                                    f"  ·  {int(hit_score*100)}% match"
                                )
                                hit_options.append(lbl)
                                hit_map[lbl] = hit_person

                            # Determine current selection index
                            current_pid  = dec.get("resolved_pid", "")
                            current_name = dec.get("resolved_name", author)
                            default_sel  = 0  # new person
                            if current_pid and current_pid != first.get("suggested_pid", ""):
                                # A real existing person was previously selected — find in options
                                for i, lbl in enumerate(hit_options):
                                    if f"[{current_pid}]" in lbl:
                                        default_sel = i
                                        break
                                if default_sel == 0 and current_pid:
                                    # Previously-selected person not in current search results;
                                    # show them as a pinned option so the selection persists
                                    pinned_lbl = f"[{current_pid}] {current_name}  ·  selected"
                                    hit_options.insert(1, pinned_lbl)
                                    hit_map[pinned_lbl] = {
                                        "PersonID": current_pid,
                                        "AuthorFullName": current_name,
                                        "OrganizationID": dec.get("org_ids", [""])[0],
                                        "OrganizationIDs": dec.get("org_ids", [""]),
                                    }
                                    default_sel = 1

                            chosen_lbl = st.selectbox(
                                f"Identity for {author}",
                                options=hit_options,
                                index=default_sel,
                                key=f"identity_new_{norm}",
                            )

                            if chosen_lbl == NEW_PERSON_LABEL:
                                dec["resolved_pid"]  = first.get("suggested_pid", "")
                                dec["resolved_name"] = author
                                dec["match_type"]    = "new"
                                dec["org_ids"]       = dec.get("org_ids") or [""]
                                st.caption(f"Will be registered as a new person · ID {dec['resolved_pid']}")
                            else:
                                chosen_person = hit_map[chosen_lbl]
                                dec["resolved_pid"]  = chosen_person["PersonID"]
                                dec["resolved_name"] = chosen_person["AuthorFullName"]
                                dec["match_type"]    = "resolved"
                                dec["org_ids"] = chosen_person.get("OrganizationIDs") or (
                                    [chosen_person["OrganizationID"]]
                                    if chosen_person.get("OrganizationID") else [""]
                                )
                                # Check immediately if this is a duplicate
                                ep = st.session_state.get("existing_pairs", set())
                                for it in items:
                                    if (chosen_person["PersonID"], it["UT"]) in ep:
                                        st.warning(
                                            f"⚠️ **Probable duplicate** — "
                                            f"{chosen_person['AuthorFullName']} (ID {chosen_person['PersonID']}) "
                                            f"is already linked to **{it['UT']}** in MyOrg. "
                                            f"If you approve this, it will be **skipped** (not re-uploaded).",
                                            icon="🔁",
                                        )
                                        break

                    with id_col2:
                        if cfg.get("allow_multi_org", True):
                            selected_labels = st.multiselect(
                                "Assign organization(s)",
                                options=list(org_map.keys()),
                                default=[
                                    label_for_org(oid)
                                    for oid in dec.get("org_ids", [""])
                                    if oid and label_for_org(oid) != org_labels[0]
                                ],
                                key=f"orgs_{norm}",
                            )
                            dec["org_ids"] = [org_map[lbl] for lbl in selected_labels] or [""]
                        else:
                            sel = st.selectbox(
                                "Assign organization",
                                options=org_labels,
                                key=f"org_{norm}",
                            )
                            dec["org_ids"] = [org_map[sel]] if sel in org_map else [""]

                    st.markdown("---")
                    ap_col, rj_col = st.columns(2)
                    with ap_col:
                        approved_val = st.checkbox(
                            "✅ **Approve & include in upload**",
                            value=dec.get("approved", mt != "new"),
                            key=f"approve_{norm}",
                            help="Check to include this entry in the upload-ready CSV"
                        )
                        if approved_val != dec.get("approved", mt != "new"):
                            dec["user_decided"] = True
                        dec["approved"] = approved_val
                        # Also mark as decided once checkbox has been rendered and interacted with
                        if f"approve_{norm}" in st.session_state:
                            dec["user_decided"] = True
                    with rj_col:
                        if not dec.get("approved", True):
                            st.markdown(
                                '<div style="background:#fdecea;border-left:4px solid #e74c3c;'
                                'padding:0.5rem 0.8rem;border-radius:0 6px 6px 0;margin-top:0.3rem;">'
                                '⛔ This entry will be <b>excluded</b> from the output.</div>',
                                unsafe_allow_html=True
                            )
                        else:
                            st.markdown(
                                '<div style="background:#eafaf1;border-left:4px solid #27ae60;'
                                'padding:0.5rem 0.8rem;border-radius:0 6px 6px 0;margin-top:0.3rem;">'
                                '✅ This entry will be <b>included</b> in the upload.</div>',
                                unsafe_allow_html=True
                            )
                    decisions[norm] = dec

        st.session_state.decisions = decisions

        st.markdown("---")

        if st.button("💾  Save Decisions & Prepare Output", type="primary",
                     use_container_width=True):
            output_rows   = []
            rejected_rows = []
            seen: set[tuple] = set()

            # Auto-confirmed exact matches
            for row in confirmed_auto:
                key = (row["PersonID"], row["UT"], row["OrganizationID"])
                if key not in seen:
                    seen.add(key)
                    output_rows.append(row)

            # User decisions
            needs_by_norm: dict[str, list] = defaultdict(list)
            for item in needs_review:
                needs_by_norm[item["norm"]].append(item)

            existing_pairs_save = st.session_state.get("existing_pairs", set())

            for norm, items in needs_by_norm.items():
                dec = decisions.get(norm)
                item_mt = items[0].get("match_type", "new")
                # New persons require explicit approval; fuzzy/initial default to approved
                default_approved = item_mt != "new"
                if not dec or not dec.get("approved", default_approved):
                    for it in items:
                        rejected_rows.append({
                            "AuthorFullName": it["AuthorFullName"],
                            "UT": it["UT"],
                            "Reason": "User rejected" if dec else "No decision made",
                        })
                    continue

                pid           = str(dec.get("resolved_pid", "")).strip()
                # Fall back to staged PID only if resolved_pid is genuinely empty
                # (i.e. user never selected anything — shouldn't happen with new UI)
                if not pid:
                    pid = str(items[0].get("suggested_pid", "")).strip()
                resolved_name = dec.get("resolved_name", items[0]["AuthorFullName"])
                org_ids       = [o for o in dec.get("org_ids", [""]) if o] or [""]

                for item in items:
                    # Check if this (PersonID, UT) is already in MyOrg
                    # Catches cases where user resolved a 'new' person to an existing PID
                    if pid and (pid, item["UT"]) in existing_pairs_save:
                        rejected_rows.append({
                            "AuthorFullName": resolved_name,
                            "UT": item["UT"],
                            "Reason": "Already in MyOrg (resolved match)",
                        })
                        continue

                    for oid in org_ids:
                        key = (pid, item["UT"], oid)
                        if key in seen:
                            rejected_rows.append({
                                "AuthorFullName": resolved_name,
                                "UT": item["UT"],
                                "Reason": "Duplicate",
                            })
                            continue
                        seen.add(key)
                        output_rows.append({
                            "PersonID":       pid,
                            "AuthorFullName": resolved_name,
                            "UT":             item["UT"],
                            "OrganizationID": oid,
                            "match_type":     dec.get("match_type", ""),
                        })

            st.session_state.output_rows   = output_rows
            st.session_state.rejected_rows = rejected_rows
            st.session_state.finalized     = True
            st.success(
                f"✅ {len(output_rows)} rows finalized "
                f"({len(rejected_rows)} rejected). Go to **Tab 3** to export."
            )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — EXPORT OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

with tab_output:
    if not st.session_state.processed:
        st.info("⬅️ Please complete Tabs 1 and 2 first.")
    elif not st.session_state.finalized:
        st.warning("⚠️ Please save decisions in **Tab 2** before exporting.")
    else:
        output_rows   = st.session_state.output_rows
        rejected_rows = st.session_state.rejected_rows
        source_file   = st.session_state.source_file
        orgs          = st.session_state.orgs
        batch_result  = st.session_state.batch_result

        st.markdown('<div class="sec-head">📤 Export Files</div>', unsafe_allow_html=True)

        st.markdown(f"""
<div class="metric-grid">
  <div class="metric-card"><div class="num num-green">{len(output_rows)}</div><div class="lbl">Upload-Ready Rows</div></div>
  <div class="metric-card"><div class="num num-orange">{len(rejected_rows)}</div><div class="lbl">Rejected / Skipped</div></div>
</div>
""", unsafe_allow_html=True)

        # ── Upload-ready CSV ─────────────────────────────────────────────────
        st.markdown("#### 1. Upload-Ready CSV")
        st.markdown("Compatible with WoS My Organization bulk import format.")

        csv_bytes = build_upload_csv(output_rows, source_file).encode("utf-8")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        st.download_button(
            label="⬇️  Download Upload-Ready CSV",
            data=csv_bytes,
            file_name=f"upload_ready_{ts}.csv",
            mime="text/csv",
            use_container_width=True,
        )

        with st.expander("Preview upload CSV", expanded=False):
            st.dataframe(pd.DataFrame(output_rows).head(30), use_container_width=True)

        st.markdown("---")

        # ── Review Excel ─────────────────────────────────────────────────────
        needs_review = batch_result.get("needs_review", [])
        if needs_review:
            st.markdown("#### 2. Review Excel (for batch workflows)")
            st.markdown("Share with library staff to fill in decisions offline.")
            if HAS_OPENPYXL:
                excel_bytes = build_review_excel(needs_review, orgs)
                st.download_button(
                    label="⬇️  Download Review Excel",
                    data=excel_bytes,
                    file_name=f"review_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            else:
                st.warning("openpyxl not installed — Excel export unavailable.")
            st.markdown("---")

        # ── Audit log ────────────────────────────────────────────────────────
        st.markdown("#### 3. Audit Log (JSON)")
        new_persons_list = batch_result.get("new_persons", [])
        if isinstance(new_persons_list, dict):
            new_persons_list = list(new_persons_list.values())

        n_exact   = len([r for r in batch_result.get("confirmed", [])
                         if r.get("match_type") == "exact"])
        n_initial = len([r for r in batch_result.get("needs_review", [])
                         if r.get("match_type") == "initial_expansion"])
        n_fuzzy   = len([r for r in batch_result.get("needs_review", [])
                         if r.get("match_type") == "fuzzy"])
        n_new     = len([r for r in batch_result.get("needs_review", [])
                         if r.get("match_type") == "new"])

        audit_json = build_audit_json(
            summary={
                "exact_matches":             n_exact,
                "initial_expansion_matches": n_initial,
                "fuzzy_matches":             n_fuzzy,
                "new_persons":               n_new,
                "finalized_records":         len(output_rows),
                "rejected_records":          len(rejected_rows),
            },
            new_persons=new_persons_list,
        )
        st.download_button(
            label="⬇️  Download Audit Log (JSON)",
            data=audit_json.encode("utf-8"),
            file_name=f"audit_{ts}.json",
            mime="application/json",
            use_container_width=True,
        )

        with st.expander("Preview audit log", expanded=False):
            st.json(json.loads(audit_json))

        st.markdown("---")

        # ── Re-import filled Excel ────────────────────────────────────────────
        st.markdown("#### 4. Re-import Filled Review Excel")
        reimport_file = st.file_uploader("Upload filled review Excel",
                                         type=["xlsx"], key="reimport")
        if reimport_file and st.button("🔄 Merge Review Decisions"):
            wb      = __import__("openpyxl").load_workbook(reimport_file)
            ws      = wb["Author Review"]
            headers = [c.value for c in ws[1]]

            def col(n):
                return headers.index(n)

            extra_rows = []
            skip_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                approved = row[col("APPROVED")]
                if approved and str(approved).strip().upper() == "YES":
                    extra_rows.append({
                        "PersonID":       str(row[col("Detected PersonID")] or ""),
                        "AuthorFullName": str(row[col("Existing Name")]      or ""),
                        "UT":             str(row[col("UT")]                  or ""),
                        "OrganizationID": str(row[col("OrganizationID")]      or ""),
                        # UT is also stored as DocumentID alias for build_upload_csv
                        "DocumentID":     str(row[col("UT")]                  or ""),
                    })
                else:
                    skip_count += 1

            merged     = output_rows + extra_rows
            merged_csv = build_upload_csv(merged, source_file).encode("utf-8")
            st.success(
                f"✅ Merged {len(merged)} rows "
                f"({len(extra_rows)} from review, {skip_count} skipped)"
            )
            st.download_button(
                "⬇️ Download Merged CSV",
                data=merged_csv,
                file_name=f"upload_ready_merged_{ts}.csv",
                mime="text/csv",
                use_container_width=True,
            )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — STATISTICS
# ══════════════════════════════════════════════════════════════════════════════

with tab_stats:
    if not st.session_state.processed:
        st.info("⬅️ Load and process data in Tab 1 to see statistics.")
    else:
        muv_pairs    = st.session_state.muv_pairs
        batch_result = st.session_state.batch_result
        person_index = st.session_state.person_index
        orgs         = st.session_state.orgs

        confirmed    = batch_result["confirmed"]
        needs_review = batch_result["needs_review"]
        new_persons  = batch_result["new_persons"]
        if isinstance(new_persons, dict):
            new_persons = list(new_persons.values())

        st.markdown('<div class="sec-head">Processing Summary</div>', unsafe_allow_html=True)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("WoS Records",           len(st.session_state.wos_records))
        c2.metric("MUV Author-Doc Pairs",  len(muv_pairs))
        c3.metric("Existing Persons",      len(person_index))
        c4.metric("Organizations",         len(orgs))

        already_up = batch_result.get("already_uploaded", [])

        c5, c6, c7, c8 = st.columns(4)
        c5.metric("Auto-Confirmed (exact)",    len(confirmed))
        c6.metric("Initial-Expansion Matches", len([r for r in needs_review if r["match_type"] == "initial_expansion"]))
        c7.metric("Fuzzy Matches",             len([r for r in needs_review if r["match_type"] == "fuzzy"]))
        c8.metric("New Persons Staged",        len(new_persons))

        already_up_exact = [r for r in already_up if r.get("match_type") != "probable_duplicate"]
        already_up_prob  = [r for r in already_up if r.get("match_type") == "probable_duplicate"]

        c9, c10, c11, _ = st.columns(4)
        c9.metric("Needs Review",        len(needs_review))
        c10.metric("Already in MyOrg",   len(already_up_exact),
                   help="Exact (PersonID, UT) pairs already in ResearcherAndDocument.csv")
        c11.metric("Probable Duplicates", len(already_up_prob),
                   help="High-confidence name matches (≥0.90) whose UT already exists for that person")

        if already_up_prob:
            st.markdown("---")
            st.markdown('<div class="sec-head">⚠ Probable Duplicates — auto-skipped</div>', unsafe_allow_html=True)
            st.caption(
                "These are high-confidence name matches (initial-expansion or fuzzy, score ≥ 0.90) "
                "where the matched PersonID already has this UT in ResearcherAndDocument.csv. "
                "They were skipped automatically to avoid re-uploading existing records."
            )
            df_prob = pd.DataFrame([{
                "WoS Name":    r.get("author_full", r.get("AuthorFullName", "")),
                "Matched To":  r.get("AuthorFullName", ""),
                "PersonID":    r.get("PersonID", ""),
                "Score":       f'{r.get("match_score", 0):.2f}',
                "UT":          r.get("UT", r.get("ut", "")),
                "OrgID":       r.get("OrganizationID", ""),
                "Reason":      r.get("Reason", ""),
            } for r in already_up_prob])
            st.dataframe(df_prob, use_container_width=True)

        if already_up_exact:
            st.markdown("---")
            st.markdown('<div class="sec-head">⏭ Already in MyOrg — exact skip</div>', unsafe_allow_html=True)
            df_up = pd.DataFrame([{
                "PersonID":   r.get("PersonID", ""),
                "Author":     r.get("AuthorFullName", ""),
                "UT":         r.get("UT", r.get("ut", "")),
                "OrgID":      r.get("OrganizationID", ""),
                "Reason":     r.get("Reason", "Already in MyOrg"),
            } for r in already_up_exact])
            st.dataframe(df_up, use_container_width=True)

        st.markdown("---")

        st.markdown('<div class="sec-head">MUV Authors in Input</div>', unsafe_allow_html=True)
        author_doc_counts: dict[str, int] = defaultdict(int)
        for p in muv_pairs:
            author_doc_counts[p["author_full"]] += 1

        if author_doc_counts:
            df_authors = pd.DataFrame([
                {"Author": k, "Documents": v}
                for k, v in sorted(author_doc_counts.items(), key=lambda x: -x[1])
            ])
            st.bar_chart(df_authors.set_index("Author")["Documents"])

        st.markdown("---")

        st.markdown('<div class="sec-head">MUV Affiliation Strings Detected</div>',
                    unsafe_allow_html=True)
        affil_counts: dict[str, int] = defaultdict(int)
        for p in muv_pairs:
            for a in p["muv_affils"]:
                affil_counts[a] += 1

        if affil_counts:
            df_affils = pd.DataFrame([
                {"Affiliation": k, "Count": v}
                for k, v in sorted(affil_counts.items(), key=lambda x: -x[1])
            ])
            st.dataframe(df_affils, use_container_width=True)

        st.markdown("---")

        st.markdown('<div class="sec-head">All MUV Author-Document Pairs</div>',
                    unsafe_allow_html=True)
        df_pairs = pd.DataFrame([{
            "Author":           p["author_full"],
            "UT":               p["ut"],
            "MUV Affiliations": " | ".join(p["muv_affils"]),
        } for p in muv_pairs])
        st.dataframe(df_pairs, use_container_width=True, height=300)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — HELP
# ══════════════════════════════════════════════════════════════════════════════

with tab_help:
    st.markdown("""
## WoS MUV Affiliation Ingestion Tool · User Guide

### Overview
This tool identifies **Medical University of Varna (MUV)**-affiliated authors from
Web of Science (WoS) exports and generates upload-ready entries for the
**WoS My Organization** module.

---

### Workflow

| Step | Tab | What to do |
|------|-----|-----------|
| 1 | Load Files | Upload WoS export CSV, ResearcherAndDocument.csv, and OrganizationHierarchy.csv |
| 2 | Review & Resolve | Verify matches, assign organizations, approve/reject entries |
| 3 | Export Output | Download upload-ready CSV, review Excel, and audit log |

---

### Input Files

**WoS Export CSV** (`input*.csv`)
- Must contain columns: `AF`, `C1`, `UT`
- Export from WoS using "Tab-delimited (Win, UTF-8)" format → save as CSV

**ResearcherAndDocument.csv**
- Download from WoS My Organization → Export
- Columns: `PersonID`, `FirstName`, `LastName`, `OrganizationID`, `DocumentID`

**OrganizationHierarchy.csv**
- Download from WoS My Organization → Settings → Org Hierarchy
- Columns: `OrganizationID`, `OrganizationName`, `ParentOrgaID`

---

### Match Types

| Badge | Meaning | Action required |
|-------|---------|-----------------|
| ✓ EXACT | Exact name match found | None — auto-confirmed |
| 🔤 INITIAL MATCH | WoS initials are compatible with a master full name (e.g. `Lazarov, N.` → `Lazarov, Nikola R.`) | Confirm or redirect to a different person |
| ⚠ AMBIGUOUS | Name similar to existing person(s) | Choose correct person or create new |
| 🆕 NEW PERSON | No match found | Verify and assign organization |

### Sibling Groups
When two WoS name variants refer to the same person (e.g. `Lazarov, N. R.` and
`Lazarov, N.`), they are shown under a **🔗 Sibling group** banner so you can
resolve them together.

---

### Configuration (sidebar)

- **Fuzzy threshold**: 0.85 recommended. Lower = more fuzzy matches surfaced.
- **MUV patterns**: Customize to catch transliteration variants or new sub-units.
- **Multi-org**: Allow assigning one author to multiple organizational units.

---

### Batch Workflow (for library teams)

1. Run processing → download **Review Excel** from Tab 3
2. Share Excel with curators to fill in `OrganizationID` and set `APPROVED = YES`
3. Return to the app → Tab 3 → upload filled Excel → download merged CSV
4. Import merged CSV into WoS My Organization
""")
