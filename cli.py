#!/usr/bin/env python3
"""
cli.py — WoS MUV Affiliation Ingestion Tool · CLI Mode
Medical University of Varna · Research Information Systems

Usage:
  python cli.py input100.csv --mode interactive
  python cli.py input100.csv --mode batch
  python cli.py input100.csv --mode batch --reimport review_filled.xlsx
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import logging
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from core import (
    DEFAULT_CONFIG, load_config,
    normalize_name, name_similarity,
    build_person_index, build_researcher_dataframe,
    parse_org_hierarchy, parse_wos_csv,
    extract_muv_author_pairs, match_person, batch_process,
    build_upload_csv, build_audit_json, build_review_excel,
    StagingDB,
)
from initial_matching import InitialAwareMatcher

# ─── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("wos_muv.cli")

# ─── Try rich for prettier output ─────────────────────────────────────────────
try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich import print as rprint
    console = Console()
    HAS_RICH = True
except ImportError:
    console = None
    HAS_RICH = False


def cprint(msg: str, style: str = ""):
    if HAS_RICH:
        console.print(msg, style=style) if style else console.print(msg)
    else:
        import re
        clean = re.sub(r"\[/?[^\]]+\]", "", msg)
        print(clean)


def banner():
    cprint("\n[bold blue]╔══════════════════════════════════════════════════╗[/bold blue]")
    cprint("[bold blue]║  WoS My Organization · MUV Ingestion Tool       ║[/bold blue]")
    cprint("[bold blue]║  Medical University of Varna                     ║[/bold blue]")
    cprint("[bold blue]╚══════════════════════════════════════════════════╝[/bold blue]\n")


# ─── Interactive helpers ──────────────────────────────────────────────────────

def prompt_fuzzy_resolve(author_full: str, candidates: list, ut: str,
                         match_subtype: str = "fuzzy") -> tuple[str, str]:
    """
    Ask user to choose between fuzzy/initial-expansion candidates or create new.
    Returns (person_id, full_name).
    """
    label = (
        "🔤 Initial-expansion match" if match_subtype == "initial_expansion"
        else "⚠ Ambiguous match"
    )
    cprint(f"\n[bold yellow]{label} for:[/bold yellow] [cyan]{author_full}[/cyan]  (UT: {ut})")
    cprint("  Possible matches in existing dataset:")
    for i, (_, person, score) in enumerate(candidates, 1):
        subtype_tag = (
            " [initial match]" if person.get("_match_subtype") == "initial_expansion" else ""
        )
        cprint(f"    {i}. {person['AuthorFullName']}{subtype_tag}  "
               f"(PersonID {person['PersonID']}, score: {score:.2f})")
    cprint(f"    {len(candidates)+1}. ➕  Create as NEW PERSON")

    while True:
        try:
            raw = input(f"  Choice [1-{len(candidates)+1}]: ").strip()
            idx = int(raw)
            if 1 <= idx <= len(candidates):
                p = candidates[idx - 1][1]
                return p["PersonID"], p["AuthorFullName"]
            elif idx == len(candidates) + 1:
                return "NEW", author_full
        except (ValueError, KeyboardInterrupt):
            cprint("  Invalid input, please try again.")


def prompt_org_selection(author_full: str, orgs: list[dict],
                         suggested_org_ids: list[str] = None,
                         multi: bool = True) -> list[str]:
    """
    Ask user to assign one or more OrganizationIDs.
    Pre-selects suggested_org_ids (from master record) if provided.
    Returns list of org ID strings.
    """
    # If we have suggested orgs from the master record, offer to use them directly
    if suggested_org_ids:
        org_names = []
        for oid in suggested_org_ids:
            match = next((o for o in orgs if o["OrganizationID"] == oid), None)
            name = match["OrganizationName"] if match else oid
            org_names.append(f"[{oid}] {name}")
        cprint(f"\n  Suggested organization(s) from master record: "
               f"[green]{', '.join(org_names)}[/green]")
        ans = input("  Use suggested org(s)? [Y/n]: ").strip().lower()
        if ans != "n":
            return suggested_org_ids

    cprint(f"\n[bold cyan]🏛  Select organization for:[/bold cyan] [green]{author_full}[/green]")

    roots    = [o for o in orgs if not o["ParentOrgaID"]]
    children = [o for o in orgs if o["ParentOrgaID"]]
    display  = roots + children

    for i, org in enumerate(display, 1):
        indent = "    " if org["ParentOrgaID"] else ""
        cprint(f"    {i:3}. {indent}[{org['OrganizationID']}] {org['OrganizationName']}")

    n = len(display)
    cprint(f"    {n+1:3}. 🔍  Search by keyword")
    cprint(f"    {n+2:3}. ✏️   Enter custom OrganizationID")
    cprint(f"      0. Skip (no organization)")

    assigned: list[str] = []

    def pick_one() -> str | None:
        while True:
            try:
                raw = input(f"  Choice [0-{n+2}]: ").strip()
                idx = int(raw)
                if idx == 0:
                    return None
                elif 1 <= idx <= n:
                    return display[idx - 1]["OrganizationID"]
                elif idx == n + 1:
                    kw   = input("  Search keyword: ").strip().lower()
                    hits = [o for o in orgs if kw in o["OrganizationName"].lower()]
                    if not hits:
                        cprint("  No matches found.")
                        continue
                    for j, o in enumerate(hits, 1):
                        cprint(f"    {j}. [{o['OrganizationID']}] {o['OrganizationName']}")
                    sub = input(f"  Select [1-{len(hits)}]: ").strip()
                    try:
                        return hits[int(sub) - 1]["OrganizationID"]
                    except (ValueError, IndexError):
                        continue
                elif idx == n + 2:
                    return input("  Custom OrganizationID: ").strip() or None
            except (ValueError, KeyboardInterrupt):
                cprint("  Invalid input.")

    org_id = pick_one()
    if org_id:
        assigned.append(org_id)

    if multi and org_id:
        while True:
            more = input("  Add another organization? [y/N]: ").strip().lower()
            if more != "y":
                break
            extra = pick_one()
            if extra and extra not in assigned:
                assigned.append(extra)

    return assigned


# ─── Interactive mode ─────────────────────────────────────────────────────────

def run_interactive(
    wos_content: str,
    researcher_content: str,
    org_content: str,
    cfg: dict,
    source_file: str,
    out_dir: str,
) -> dict:
    os.makedirs(out_dir, exist_ok=True)
    ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
    db        = StagingDB(os.path.join(out_dir, cfg.get("db_path", "staging.db")))

    person_index, max_pid = build_person_index(researcher_content)
    orgs      = parse_org_hierarchy(org_content)
    records   = parse_wos_csv(wos_content)
    muv_pairs = extract_muv_author_pairs(records, cfg)

    # Build InitialAwareMatcher so interactive mode benefits from initial-expansion
    researcher_df  = build_researcher_dataframe(researcher_content)
    threshold      = float(cfg.get("fuzzy_threshold", 0.85))
    initial_matcher = (
        InitialAwareMatcher(researcher_df, fuzzy_threshold=threshold)
        if researcher_df is not None else None
    )

    # Always start above the highest existing PersonID — never re-use IDs
    start_pid   = max_pid + 1
    pid_counter = start_pid
    multi_org   = cfg.get("allow_multi_org", True)

    output_rows: list[dict]        = []
    rejected: list[dict]           = []
    new_persons_created: dict[str, dict] = {}
    seen_pairs: set[tuple]         = set()

    author_groups: dict[str, list[dict]] = defaultdict(list)
    for pair in muv_pairs:
        norm = normalize_name(pair["author_full"])
        author_groups[norm].append(pair)

    total = len(author_groups)
    cprint(f"\n[bold]Found {len(muv_pairs)} MUV author-document pairs "
           f"across {total} unique authors.[/bold]")

    for auth_idx, (norm, pairs) in enumerate(author_groups.items(), 1):
        author_full = pairs[0]["author_full"]
        cprint(f"\n[dim]─── Author {auth_idx}/{total} ───[/dim]")

        match_type, candidates = match_person(
            author_full, person_index,
            {"fuzzy_threshold": threshold},
            initial_matcher,
        )
        match_subtype = (
            candidates[0][1].get("_match_subtype", "fuzzy")
            if candidates else "fuzzy"
        )

        # Resolve identity
        if match_type == "exact":
            person        = candidates[0][1]
            pid           = person["PersonID"]
            resolved_name = person["AuthorFullName"]
            suggested_org_ids = person.get("OrganizationIDs") or (
                [person["OrganizationID"]] if person.get("OrganizationID") else []
            )
            cprint(f"  [green]✓ Exact match:[/green] {resolved_name} (PersonID {pid})")
            db.log_decision(pid, "exact_match", author_full)

        elif match_type == "fuzzy":
            pid_raw, resolved_name = prompt_fuzzy_resolve(
                author_full, candidates, pairs[0]["ut"], match_subtype
            )
            if pid_raw == "NEW":
                match_type = "new"
                pid = None
                suggested_org_ids = []
            else:
                pid = pid_raw
                # Use org IDs from the chosen candidate
                chosen = next(
                    (c[1] for c in candidates if c[1]["PersonID"] == pid), {}
                )
                suggested_org_ids = chosen.get("OrganizationIDs") or (
                    [chosen["OrganizationID"]] if chosen.get("OrganizationID") else []
                )
                db.log_decision(pid, "fuzzy_resolved",
                                f"Chose {pid} for: {author_full}")

        if match_type == "new":
            if norm in new_persons_created:
                entry         = new_persons_created[norm]
                pid           = entry["PersonID"]
                resolved_name = entry["AuthorFullName"]
            else:
                pid           = str(pid_counter)
                pid_counter  += 1
                resolved_name = author_full
                new_persons_created[norm] = {
                    "PersonID": pid, "AuthorFullName": author_full
                }
                db.upsert_person(pid, author_full, norm, is_new=True)
                cprint(f"  [bold green]➕ NEW PERSON:[/bold green] "
                       f"{author_full} → PersonID {pid}")
                db.log_decision(pid, "new_person", author_full)
            suggested_org_ids = []

        # Org assignment
        org_assigned_for_author: list[str] | None = None

        for pair in pairs:
            ut         = pair["ut"]
            muv_affils = pair["muv_affils"]
            cprint(f"  [dim]UT: {ut} | MUV affils: {'; '.join(muv_affils[:2])}[/dim]")

            if match_type == "exact":
                org_ids = suggested_org_ids or prompt_org_selection(
                    author_full, orgs, suggested_org_ids=[], multi=multi_org
                )
            else:
                if org_assigned_for_author is None:
                    cprint(f"  Affiliations detected: {'; '.join(muv_affils)}")
                    org_assigned_for_author = prompt_org_selection(
                        author_full, orgs,
                        suggested_org_ids=suggested_org_ids,
                        multi=multi_org,
                    )
                org_ids = org_assigned_for_author

            if not org_ids:
                org_ids = [""]

            for org_id in org_ids:
                key = (pid, ut, org_id)
                if key in seen_pairs:
                    db.log_rejected(author_full, ut, f"Duplicate pair {key}")
                    rejected.append({
                        "AuthorFullName": author_full,
                        "UT": ut,
                        "Reason": "Duplicate",
                    })
                    continue
                seen_pairs.add(key)
                output_rows.append({
                    "PersonID":       pid,
                    "AuthorFullName": resolved_name,
                    "UT":             ut,
                    "OrganizationID": org_id,
                    "match_type":     match_type,
                })
                db.log_decision(pid, "org_assigned", f"UT={ut} OrgID={org_id}")

    # ── Write outputs ─────────────────────────────────────────────────────────
    csv_path = os.path.join(out_dir, f"upload_ready_{ts}.csv")
    Path(csv_path).write_text(
        build_upload_csv(output_rows, source_file), encoding="utf-8"
    )

    new_persons_list = [
        {"PersonID": v["PersonID"], "AuthorFullName": v["AuthorFullName"]}
        for v in new_persons_created.values()
    ]
    n_exact   = sum(1 for r in output_rows if r.get("match_type") == "exact")
    n_fuzzy   = sum(1 for r in output_rows if r.get("match_type") in ("fuzzy", "resolved"))
    n_initial = sum(1 for r in output_rows if r.get("match_type") == "initial_expansion")
    n_new     = len(new_persons_created)

    audit_path = os.path.join(out_dir, f"audit_{ts}.json")
    Path(audit_path).write_text(
        build_audit_json(
            summary={
                "exact_matches":             n_exact,
                "initial_expansion_matches": n_initial,
                "fuzzy_matches":             n_fuzzy,
                "new_persons":               n_new,
                "finalized_records":         len(output_rows),
                "rejected_records":          len(rejected),
            },
            new_persons=new_persons_list,
        ),
        encoding="utf-8",
    )

    cprint(f"\n[bold green]{'═'*50}[/bold green]")
    cprint("[bold green]✅ Processing complete![/bold green]")
    cprint(f"  Records in input     : {len(records)}")
    cprint(f"  MUV author-doc pairs : {len(muv_pairs)}")
    cprint(f"  Unique authors       : {len(author_groups)}")
    cprint(f"  New persons created  : {len(new_persons_created)}")
    cprint(f"  Output rows          : {len(output_rows)}")
    cprint(f"  Upload CSV           : {csv_path}")
    cprint(f"  Audit log            : {audit_path}")

    return {
        "records":      len(records),
        "muv_pairs":    len(muv_pairs),
        "new_persons":  len(new_persons_created),
        "output_rows":  len(output_rows),
        "csv_path":     csv_path,
        "audit_path":   audit_path,
    }


# ─── Batch mode ───────────────────────────────────────────────────────────────

def run_batch(
    wos_content: str,
    researcher_content: str,
    org_content: str,
    cfg: dict,
    source_file: str,
    out_dir: str,
) -> dict:
    os.makedirs(out_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    db = StagingDB(os.path.join(out_dir, cfg.get("db_path", "staging.db")))

    person_index, max_pid = build_person_index(researcher_content)
    orgs      = parse_org_hierarchy(org_content)
    records   = parse_wos_csv(wos_content)
    muv_pairs = extract_muv_author_pairs(records, cfg)

    # Always start above the highest existing PersonID — never re-use IDs
    start_pid = max_pid + 1

    # ── KEY FIX: pass researcher_content so InitialAwareMatcher gets built ───
    result = batch_process(
        muv_pairs, person_index, orgs, cfg, start_pid,
        researcher_csv_content=researcher_content,
    )
    confirmed    = result["confirmed"]
    needs_review = result["needs_review"]
    new_persons  = result["new_persons"]

    # Write confirmed rows to upload-ready CSV
    csv_path = os.path.join(out_dir, f"upload_ready_{ts}.csv")
    Path(csv_path).write_text(
        build_upload_csv(confirmed, source_file), encoding="utf-8"
    )

    # Write review Excel
    excel_path = None
    try:
        excel_bytes = build_review_excel(needs_review, orgs)
        excel_path  = os.path.join(out_dir, f"review_{ts}.xlsx")
        Path(excel_path).write_bytes(excel_bytes)
        cprint(f"  Review Excel         : {excel_path}")
    except ImportError:
        cprint("  [yellow]openpyxl not available — review Excel skipped.[/yellow]")

    # Audit log
    if isinstance(new_persons, dict):
        new_persons = list(new_persons.values())
    new_list = [
        {"PersonID": v["PersonID"], "AuthorFullName": v["AuthorFullName"]}
        for v in new_persons
    ]

    n_exact   = len(confirmed)
    n_initial = sum(1 for r in needs_review if r.get("match_type") == "initial_expansion")
    n_fuzzy   = sum(1 for r in needs_review if r.get("match_type") == "fuzzy")
    n_new     = sum(1 for r in needs_review if r.get("match_type") == "new")

    audit_path = os.path.join(out_dir, f"audit_{ts}.json")
    Path(audit_path).write_text(
        build_audit_json(
            summary={
                "exact_matches":             n_exact,
                "initial_expansion_matches": n_initial,
                "fuzzy_matches":             n_fuzzy,
                "new_persons":               n_new,
                "finalized_records":         n_exact,
                "rejected_records":          0,
            },
            new_persons=new_list,
        ),
        encoding="utf-8",
    )

    cprint(f"\n[bold green]✅ Batch processing complete![/bold green]")
    cprint(f"  Records in input     : {len(records)}")
    cprint(f"  MUV author-doc pairs : {len(muv_pairs)}")
    cprint(f"  Auto-confirmed rows  : {len(confirmed)}")
    cprint(f"  Initial-exp matches  : {n_initial}")
    cprint(f"  Fuzzy matches        : {n_fuzzy}")
    cprint(f"  New persons staged   : {n_new}")
    cprint(f"  Upload CSV           : {csv_path}")
    cprint(f"  Audit log            : {audit_path}")

    return {
        "records":      len(records),
        "muv_pairs":    len(muv_pairs),
        "confirmed":    len(confirmed),
        "needs_review": len(needs_review),
        "new_persons":  n_new,
        "csv_path":     csv_path,
        "excel_path":   excel_path,
        "audit_path":   audit_path,
    }


def reimport_decisions(
    excel_path: str,
    confirmed_csv: str,
    source_file: str,
    out_dir: str,
) -> dict:
    """Re-import user-validated review Excel and merge with auto-confirmed CSV."""
    import openpyxl
    wb      = openpyxl.load_workbook(excel_path)
    # Sheet is named "Author Review" (set by build_review_excel in core.py)
    ws      = wb["Author Review"]
    headers = [cell.value for cell in ws[1]]

    def col(name):
        return headers.index(name)

    extra_rows = []
    skipped    = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col("APPROVED")] and str(row[col("APPROVED")]).strip().upper() == "YES":
            extra_rows.append({
                "PersonID":       str(row[col("Detected PersonID")] or ""),
                "AuthorFullName": str(row[col("Existing Name")]      or ""),
                "UT":             str(row[col("UT")]                  or ""),
                "OrganizationID": str(row[col("OrganizationID")]      or ""),
            })
        else:
            skipped += 1

    # Read existing confirmed CSV and parse it back
    merged_rows: list[dict] = []
    if confirmed_csv and os.path.exists(confirmed_csv):
        reader = csv.DictReader(
            io.StringIO(Path(confirmed_csv).read_text(encoding="utf-8"))
        )
        for r in reader:
            # The confirmed CSV is already in MyOrg format (FirstName/LastName/DocumentID)
            # Reconstruct AuthorFullName so build_upload_csv can re-split it
            merged_rows.append({
                "PersonID":       r.get("PersonID", ""),
                "AuthorFullName": f"{r.get('LastName','')}, {r.get('FirstName','')}".strip(", "),
                "UT":             r.get("DocumentID", ""),
                "OrganizationID": r.get("OrganizationID", ""),
            })

    merged_rows.extend(extra_rows)

    ts           = datetime.now().strftime("%Y%m%d_%H%M%S")
    merged_path  = os.path.join(out_dir, f"upload_ready_merged_{ts}.csv")
    Path(merged_path).write_text(
        build_upload_csv(merged_rows, source_file), encoding="utf-8"
    )

    cprint(f"  Merged rows          : {len(merged_rows)}")
    cprint(f"  Approved from review : {len(extra_rows)}")
    cprint(f"  Skipped (NOT YES)    : {skipped}")
    cprint(f"  Merged CSV           : {merged_path}")
    return {"merged_path": merged_path, "merged_rows": len(merged_rows)}


# ─── Entry Point ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="WoS My Organization · MUV Affiliation Ingestion Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompt for each ambiguous case):
  python cli.py input100.csv

  # Batch mode (auto-confirm exact matches, export review Excel):
  python cli.py input100.csv --mode batch

  # Re-import filled review Excel after batch mode:
  python cli.py input100.csv --mode batch --reimport output/review_*.xlsx

  # Custom file paths:
  python cli.py data/wos_export.csv \\
      --researcher data/ResearcherAndDocument.csv \\
      --orgs data/OrganizationHierarchy.csv \\
      --config config.json --out output/
        """
    )
    parser.add_argument("input",         help="WoS export CSV file")
    parser.add_argument("--researcher",  default="ResearcherAndDocument.csv",
                        help="Existing ResearcherAndDocument.csv")
    parser.add_argument("--orgs",        default="OrganizationHierarchy.csv",
                        help="OrganizationHierarchy.csv")
    parser.add_argument("--config",      default="config.json",
                        help="JSON config file")
    parser.add_argument("--mode",        choices=["interactive", "batch"],
                        default="interactive",
                        help="Processing mode (default: interactive)")
    parser.add_argument("--out",         default="output",
                        help="Output directory (default: output/)")
    parser.add_argument("--reimport",    default=None,
                        help="Path to filled review Excel to re-import (batch mode only)")
    parser.add_argument("--confirmed-csv", default=None,
                        help="Auto-confirmed CSV from previous batch run (for reimport merge)")

    args = parser.parse_args()
    banner()

    cfg = load_config(args.config)

    def read_file(path: str) -> str:
        if not os.path.exists(path):
            cprint(f"[bold red]Error:[/bold red] File not found: {path}")
            sys.exit(1)
        return Path(path).read_text(encoding="utf-8-sig")

    wos_content        = read_file(args.input)
    researcher_content = read_file(args.researcher)
    org_content        = read_file(args.orgs)
    source_file        = os.path.basename(args.input)

    if args.reimport:
        if not args.confirmed_csv:
            cprint("[yellow]Note: --confirmed-csv not provided; reimport-only mode.[/yellow]")
        cprint(f"\n[bold]Re-importing review decisions from:[/bold] {args.reimport}")
        reimport_decisions(
            args.reimport, args.confirmed_csv or "",
            source_file, args.out,
        )
        return

    if args.mode == "interactive":
        run_interactive(
            wos_content, researcher_content, org_content,
            cfg, source_file, args.out,
        )
    else:
        run_batch(
            wos_content, researcher_content, org_content,
            cfg, source_file, args.out,
        )


if __name__ == "__main__":
    main()
